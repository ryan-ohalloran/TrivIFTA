#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import numpy as np
import pandas as pd
from datetime import date, time
from openpyxl import load_workbook
from typing import Dict, Any, List
import io

class IftaData:
    """
    Class to hold relevant IFTA data for each unique VIN
    """
    def __init__(self, vin: str) -> None:

        self.vin = vin
        self.data = []
    
    def add_entry(self, reading_date: date, reading_time: time, odometer: int, jurisdiction: str) -> None:
        self.data.append({
            'ReadingDate': reading_date,
            'ReadingTime': reading_time,
            'Odometer': odometer,
            'Jurisdiction': jurisdiction
        })

class IftaDataCollection(Dict[str, IftaData]):
    """
    Class to hold data for all VINs for which IFTA reporting is needed -- behaves like Dict[str, IftaData]
    """
    def __init__(self) -> None:
        super().__init__()
        self.total_vehicles = 0
        self.num_nonmoving_vehicles = 0

    def add_ifta_data(self, vin: str) -> None:
        if vin not in self:
            self[vin] = IftaData(vin)
    
    def get_ifta_data(self, vin: str) -> IftaData:
        ifta_data = self.get(vin, None)
        if not ifta_data:
            raise Exception(f'IftaData not found for VIN: {vin}')
        return ifta_data

    def to_dataframe(self, remove_nonmoving_vehicles: bool = False) -> pd.DataFrame:
        """
        Convert data for all VINs to a single dataframe object,
            sorted by VIN, ReadingDate, ReadingTime
        """
        all_entries = []

        # Iterate over VINs in the collection
        for vin, ifta_data in self.items():
            for entry in ifta_data.data:
                all_entries.append({
                    'VIN': vin,
                    'ReadingDate': entry['ReadingDate'],
                    'ReadingTime': entry['ReadingTime'],
                    'Odometer': entry['Odometer'],
                    'Jurisdiction': entry['Jurisdiction']
                })

        # Create a DataFrame from the accumulated entries
        df = pd.DataFrame(all_entries)

        # Sort the DataFrame by VIN, ReadingDate, and ReadingTime
        df.sort_values(by=['VIN', 'ReadingDate', 'ReadingTime'], inplace=True)

        # Count number of total vehicles (i.e. number of unique vins)
        self.total_vehicles = len(df['VIN'].unique())

        # Retrieve list of nonmoving vehicles
        nonmoving_vehicles = self.get_nonmoving_vehicles(df)
        self.num_nonmoving_vehicles = len(nonmoving_vehicles)

        # Remove entries for vins where the odometer reading does not change
        if remove_nonmoving_vehicles:
            self.remove_unchanged(df, nonmoving_vehicles)

        # Return the dataframe
        return df
    
    def get_nonmoving_vehicles(self, df: pd.DataFrame) -> List[str]:
        """
        Retrieve all entries where the odometer reading does not change
        """
        # Group the DataFrame by 'VIN'
        grouped = df.groupby('VIN')

        # Find the vins that have exactly two entries and the odometer reading and jurisdiction are the same for both
        vins_to_remove = [vin for vin, group in grouped if len(group) == 2 and group['Odometer'].nunique() == 1 and group['Jurisdiction'].nunique() == 1]

        return vins_to_remove

    def remove_unchanged(self, df: pd.DataFrame, vins_to_remove: List[str]) -> None:
        """
        Remove entries for vins where the odometer reading does not change
        Returns the number of entries removed
        """
        # Remove the entries for these vins
        df.drop(df[df['VIN'].isin(vins_to_remove)].index, inplace=True)

    def export_data(self, output_file: str) -> None:
        """
        Export data for all VINs to a CSV file
        """
        df = self.to_dataframe()
        df.to_csv(output_file, index=False, encoding='utf-8')

class FleetDataFrame(pd.DataFrame):
    """
    Class to encapsulate operations on a DataFrame for fleet data
    """
    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)

    def reduce_df(self) -> None:
        """
        Reduce the DataFrame to specific columns
        """
        columns_to_keep = ['FuelTaxVin', 'FuelTaxEnterTime', 'FuelTaxExitTime', 'FuelTaxJurisdiction', 'FuelTaxEnterOdometer', 'FuelTaxExitOdometer']
        self[self.columns.intersection(columns_to_keep)]
    
    def split_date_time(self) -> None:
        """
        Extract date and time from FuelTaxEnterTime and split into separate columns
        """
        self['FuelTaxEnterTime'] = pd.to_datetime(self['FuelTaxEnterTime']).dt.floor('S')
        self['FuelTaxExitTime'] = pd.to_datetime(self['FuelTaxExitTime']).dt.floor('S')
        self['EnterReadingDate'] = self['FuelTaxEnterTime'].dt.date
        self['EnterReadingTime'] = self['FuelTaxEnterTime'].dt.time
        self['ExitReadingTime'] = self['FuelTaxExitTime'].dt.time
        self.drop(columns=['FuelTaxEnterTime', 'FuelTaxExitTime'])

class FileManager:
    """
    Class for handling file-related operations
    """
    @staticmethod
    def find_header_row(input_data: bytes, header: str) -> int:
        """
        Find the row number of the header row in an Excel file
        input_data: the contents of the input file
        header: the header row to use to find the data
        rtype: int
        """
        wb = load_workbook(filename=io.BytesIO(input_data), read_only=True)
        sheet = wb['Data']

        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            if header in row:
                return row_index
        return None

    @staticmethod
    def read_file(input_file: Any, data_type: str = 'path') -> pd.DataFrame:
        """
        Read data from an Excel file and return a Pandas DataFrame
        input_file: path to the input file or the contents of the input file
        data_type: 'path' if input_file is a path, 'bytes' if input_file is the contents of the file
        rtype: Pandas DataFrame
        """
        # If the input file is a path, read the data from the file
        if data_type == 'path':
            with open(input_file, 'rb') as f:
                input_file = f.read()
        elif data_type != 'bytes':
            raise ValueError(f"Unknown data_type: {data_type}")

        # Rest of the logic for finding the header row and reading data
        header_row = FileManager.find_header_row(input_file, 'DeviceName')
        if header_row is None:
            raise ValueError(f"Header 'DeviceName' not found in data")

        df = pd.read_excel(io.BytesIO(input_file), sheet_name='Data', skiprows=header_row)
        return df


class FuelTaxProcessor:
    """
    Class to coordinate operations of reading data, processing it, and exporting the results
    """
    def __init__(self, input_file: Any, data_type: str = 'path') -> None:
        self.file_manager = FileManager()
        self.fleet_dataframe = FleetDataFrame()
        self.input_file = input_file
        self.data_type = data_type

    @staticmethod
    def to_ifta_data_collection(df: pd.DataFrame) -> IftaDataCollection:
        """
        Transform data from the input file into the desired output format
        """
        # Create a dictionary to hold the IftaData objects
        ifta_data_collection = IftaDataCollection()

        for _, row in df.iterrows():
            vin = str(row['FuelTaxVin'])
            # skip rows with potentially empty VINs
            if vin in ('nan', 'None', ''):
                continue
            enter_reading_date = row['EnterReadingDate']
            enter_reading_time = row['EnterReadingTime']
            exit_reading_time = row['ExitReadingTime']
            try:
                enter_odometer = round(row['FuelTaxEnterOdometer'])  # round to nearest whole number
            except ValueError:
                enter_odometer = None  # or some other default value
            try:
                exit_odometer = round(row['FuelTaxExitOdometer'])
            except ValueError:
                exit_odometer = None
            jurisdiction = row['FuelTaxJurisdiction']

            # if jurisdiction is empty, skip this row
            if jurisdiction in ('nan', None, '', ' '):
                continue

            # Add this VIN to the IftaDataCollection if it doesn't already exist
            ifta_data_collection.add_ifta_data(vin)

            # Get the IftaData object for this VIN
            ifta_data = ifta_data_collection.get_ifta_data(vin)

            # Add vin entry for enter time using enter odometer reading
            ifta_data.add_entry(enter_reading_date, enter_reading_time, enter_odometer, jurisdiction)

            # Only add exit time if it is the last entry for this VIN on this day
            #   (i.e. if the exit time is 00:00, then it is the last entry for the day)
            if exit_reading_time == time(0, 0):
                # Change time to 23:59 to suit IFTA requirements
                exit_reading_time = time(23, 59)
                # In this case, use exit_reading_time and exit_odometer
                ifta_data.add_entry(enter_reading_date, exit_reading_time, exit_odometer, jurisdiction)

        return ifta_data_collection

    def process_data(self) -> IftaDataCollection:
        """
        Process data using FleetDataFrame and FileManager
        """
        # Read data from the file using FileManager
        df = self.file_manager.read_file(self.input_file, self.data_type)

        # Initialize FleetDataFrame with the read data
        self.fleet_dataframe = FleetDataFrame(df)

        # Reduce and split the FleetDataFrame
        self.fleet_dataframe.reduce_df()
        self.fleet_dataframe.split_date_time()

        # Transform FleetDataFrame into IftaDataCollection
        ifta_data_collection = self.to_ifta_data_collection(self.fleet_dataframe)

        return ifta_data_collection